from docxtpl import DocxTemplate, RichText, InlineImage
from docx.shared import Inches, Mm
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors
from copy import copy
from mailmerge import MailMerge
# with MailMerge('Site1_Port_Matrix_Table.xlsx') as document

wb = load_workbook('Site1.xlsx')
sheet = wb['Sheet1']
max_col = sheet.max_row
sheet.delete_rows(sheet.min_row, 1)
for i in range (1, max_col):
    template = 'MOP_DenHaag_V1.0.docx'
    document1 = MailMerge(template)
    document1.merge(Migration_ID = str(sheet.cell(row  2, column = 1).value),
    Group_ID = str(sheet.cell(row = 3 , column = 2).value),
    Day = str(sheet.cell(row = 4, column = 3).value),
    PEER = str(sheet.cell(row = i, column = 4).value),
    Port = str(sheet.cell(row = i, column = 5).value),
    New_Port = str(sheet.cell(row = i, column = 6).value)
)
    document1.write('excel2word'+str(sheet.cell(row = i, column = 1).value)+'.docx')
    # document1.save("MOP_Site1_V2.0.docx")
    # document1.save('MOPTEST.docx')


